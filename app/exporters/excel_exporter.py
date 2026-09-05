from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings


def _cell_value(v: Any) -> Any:
    """Convert values that openpyxl cannot write (list/dict/tuple) to string."""
    if v is None:
        return ""
    if isinstance(v, (list, tuple, set)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        return str(v)
    return v


class ExcelExporter:
    def export(self, data: dict) -> Path:
        wb = Workbook()
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="A0A0A0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(ws):
            for cell in ws[1]:
                cell.font = bold
                cell.fill = fill
                cell.border = border

        # ---------- Sheet 1: Report Info ----------
        ws = wb.active
        ws.title = "Report Info"
        ws.append(["FIELD", "VALUE"])
        style_header(ws)

        base = {
            "document_id": data.get("document_id"),
            "filename": data.get("filename"),
            "page_count": data.get("page_count"),
            "processing_engine": data.get("processing_engine"),
            "overall_confidence": data.get("overall_confidence"),
            "processed_at": data.get("processed_at"),
        }
        for k, v in (data.get("header") or {}).items():
            base[f"header.{k}"] = v
        for k, v in (data.get("footer") or {}).items():
            base[f"footer.{k}"] = v

        for k, v in base.items():
            ws.append([k, _cell_value(v)])
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 90

        # ---------- Sheet 2: Employee List ----------
        emp = wb.create_sheet("Employee List")
        headers = [
            "No",
            "Full Name",
            "Staff ID",
            "Department",
            "Attendance (hours)",
            "Exam Pass",
            "Exam Fail",
            "Discipline",
            "Course Result",
            "Certificate No",
            "Remark",
            "Source Page",
            "Confidence",
        ]
        emp.append(headers)
        style_header(emp)
        for row in data.get("employees") or []:
            emp.append(
                [
                    _cell_value(row.get("no")),
                    _cell_value(row.get("full_name", "")),
                    _cell_value(row.get("staff_id", "")),
                    _cell_value(row.get("department", "")),
                    _cell_value(row.get("attendance_hours")),
                    _cell_value(row.get("exam_pass")),
                    _cell_value(row.get("exam_fail")),
                    _cell_value(row.get("discipline_status", "")),
                    _cell_value(row.get("course_result", "")),
                    _cell_value(row.get("certificate_no", "")),
                    _cell_value(row.get("remark", "")),
                    _cell_value(row.get("source_page")),
                    _cell_value(row.get("confidence")),
                ]
            )
        for col in range(1, len(headers) + 1):
            emp.column_dimensions[get_column_letter(col)].width = 16
        emp.column_dimensions["B"].width = 28

        # ---------- Sheet 3: Error Review ----------
        review = wb.create_sheet("Error Review")
        review.append(["Page", "Field", "Severity", "Message"])
        style_header(review)
        for issue in data.get("issues") or []:
            review.append(
                [
                    _cell_value(issue.get("page")),
                    _cell_value(issue.get("field")),
                    _cell_value(issue.get("severity")),
                    _cell_value(issue.get("message")),
                ]
            )
        for col in range(1, 5):
            review.column_dimensions[get_column_letter(col)].width = 28
        review.column_dimensions["D"].width = 70

        # ---------- Sheet 4: Page Analysis ----------
        pages = wb.create_sheet("Page Analysis")
        pages.append(
            ["Page", "Type", "Confidence", "Native Text", "Form No", "Keywords", "Orientation"]
        )
        style_header(pages)
        for p in data.get("pages") or []:
            pages.append(
                [
                    _cell_value(p.get("page")),
                    _cell_value(p.get("classification")),
                    _cell_value(p.get("confidence")),
                    _cell_value(p.get("has_native_text")),
                    _cell_value(p.get("detected_form_number")),
                    _cell_value(", ".join(p.get("keywords") or [])),
                    _cell_value(p.get("orientation", 0)),
                ]
            )
        for col in range(1, 8):
            pages.column_dimensions[get_column_letter(col)].width = 18

        # ---------- Sheet 5: Document Groups ----------
        groups = wb.create_sheet("Document Groups")
        groups.append(
            ["Group ID", "Type", "Pages", "First Page", "Last Page", "Confidence", "Continuation"]
        )
        style_header(groups)
        for g in data.get("groups") or []:
            groups.append(
                [
                    _cell_value(g.get("group_id")),
                    _cell_value(g.get("type")),
                    _cell_value(", ".join(map(str, g.get("pages") or []))),
                    _cell_value(g.get("first_page")),
                    _cell_value(g.get("last_page")),
                    _cell_value(g.get("confidence")),
                    _cell_value(g.get("continuation")),
                ]
            )
        for col in range(1, 8):
            groups.column_dimensions[get_column_letter(col)].width = 16

        filename = f"export_{data['document_id'][:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out = settings.export_dir / filename
        wb.save(out)
        return out
