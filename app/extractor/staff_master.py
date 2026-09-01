"""
Staff master lookup – tra Full name chuẩn theo Staff ID.
Giảm phụ thuộc OCR tên tiếng Việt.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

# Cache trong process
_cache: Dict[str, str] = {}
_loaded_path: Optional[str] = None


def _normalize_staff_id(staff_id: str) -> str:
    if not staff_id:
        return ""
    sid = re.sub(r"\s+", "", str(staff_id)).upper()
    # Chuẩn hóa VAE##### 
    m = re.search(r"(VAE\d{5,})", sid)
    return m.group(1) if m else sid


def load_staff_master(excel_path: Path | str) -> Dict[str, str]:
    """
    Đọc Excel 2 cột: Full Name | Staff ID (thứ tự cột linh hoạt).
    Returns: { staff_id_upper: full_name }
    """
    global _cache, _loaded_path
    path = Path(excel_path)
    if not path.exists():
        _cache = {}
        _loaded_path = None
        return _cache

    # Dùng cache nếu cùng file
    if _loaded_path == str(path.resolve()) and _cache:
        return _cache

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        _cache = {}
        return _cache

    # Xác định cột Name / Staff ID từ header hoặc dữ liệu
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_col, id_col = 0, 1

    for i, h in enumerate(header):
        if any(k in h for k in ("name", "họ", "ten", "tên", "full")):
            name_col = i
        if any(k in h for k in ("staff", "id", "mã", "mae", "vae", "code")):
            id_col = i

    mapping: Dict[str, str] = {}
    start = 1 if any(header) else 0
    for row in rows[start:]:
        if not row or len(row) <= max(name_col, id_col):
            continue
        name = row[name_col]
        sid = row[id_col]
        if name is None or sid is None:
            continue
        name = str(name).strip()
        sid = _normalize_staff_id(str(sid))
        if not sid or not name:
            continue
        # Giữ bản đầu tiên nếu trùng ID
        if sid not in mapping:
            mapping[sid] = name

    _cache = mapping
    _loaded_path = str(path.resolve())
    return mapping


def lookup_name(staff_id: str, master: Dict[str, str] | None = None) -> Optional[str]:
    """Trả về Full name chuẩn hoặc None nếu không có trong master."""
    sid = _normalize_staff_id(staff_id)
    if not sid:
        return None
    data = master if master is not None else _cache
    return data.get(sid)


def default_master_path() -> Path:
    """masters/staff_ref.xlsx cạnh project root."""
    return Path(__file__).resolve().parents[2] / "masters" / "staff_ref.xlsx"


def ensure_master_loaded(path: Path | str | None = None) -> Dict[str, str]:
    p = Path(path) if path else default_master_path()
    return load_staff_master(p)
