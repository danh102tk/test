"""
Tra cứu Full Name từ file Excel staff_ref (cột Full Name + Staff ID).
Dùng khi OCR tên tiếng Việt không chuẩn (Paddle).
"""
from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

_CACHE: dict[str, str] | None = None
_CACHE_PATH: str | None = None


def _normalize_staff_id(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def load_staff_ref(path: Path | None = None) -> dict[str, str]:
    """
    Đọc Excel staff_ref → dict[staff_id] = full_name.
    Tìm cột linh hoạt: Staff ID / StaffID / Mã NV, Full Name / FullName / Họ tên.
    """
    global _CACHE, _CACHE_PATH

    if path is None:
        from app.core.config import settings
        path = settings.staff_ref_path

    path = Path(path) if path else None
    if not path or not path.exists():
        logger.info("staff_ref not found: %s", path)
        return {}

    key = str(path.resolve())
    if _CACHE is not None and _CACHE_PATH == key:
        return _CACHE

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl missing – cannot load staff_ref")
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    # Tìm header
    header = [str(c or "").strip().lower() for c in rows[0]]
    id_col = name_col = None
    for i, h in enumerate(header):
        if h in {"staff id", "staffid", "staff_id", "mã nv", "ma nv", "id"}:
            id_col = i
        if h in {"full name", "fullname", "họ tên", "ho ten", "họ và tên", "hoten"}:
            name_col = i
    # fallback: 2 cột đầu Name | ID hoặc ID | Name
    if id_col is None or name_col is None:
        if len(header) >= 2:
            name_col = name_col if name_col is not None else 0
            id_col = id_col if id_col is not None else 1

    mapping: dict[str, str] = {}
    for row in rows[1:]:
        if not row or id_col >= len(row) or name_col >= len(row):
            continue
        sid = _normalize_staff_id(str(row[id_col] or ""))
        name = str(row[name_col] or "").strip()
        if sid and name:
            mapping[sid] = name

    _CACHE = mapping
    _CACHE_PATH = key
    logger.info("Loaded staff_ref: %s entries from %s", len(mapping), path)
    return mapping


def lookup_name(staff_id: str, ref: dict[str, str] | None = None) -> str | None:
    if ref is None:
        ref = load_staff_ref()
    sid = _normalize_staff_id(staff_id)
    return ref.get(sid)
