"""
EXCEL ENGINE
Sheet 1: Report Info
Sheet 2: Employee List
Sheet 3: Error / Review (low confidence + flagged items)
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .extractor.form8014 import Form8014Result


def create_excel(result: Form8014Result, export_dir: Path, doc_id: str) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    # ------------------------------------------------------------------
    # Sheet 1 - Report Info
    # ------------------------------------------------------------------
    ws1 = wb.create_sheet("Report Info", 0)
    ws1["A1"] = "FIELD"
    ws1["B1"] = "VALUE"
    ws1["A1"].font = header_font
    ws1["B1"].font = header_font
    ws1["A1"].fill = header_fill
    ws1["B1"].fill = header_fill

    info_rows = [
        ("Form Code", result.form_code),
        ("Title", result.title),
        ("Code", result.code),
        ("Duration From", result.duration_from),
        ("Duration To", result.duration_to),
        ("Location", result.location),
        ("Total Participants", result.total_participants),
        ("Training Hours", result.training_hours),
        ("Total Training Chapters", result.total_training_chapters),
        ("Total Exam Chapters", result.total_exam_chapters),
        ("Dispatch No", result.dispatch_no),
        ("Issue Date", result.issue_date),
        ("Prepared by", result.prepared_by),
        ("Checked by", result.checked_by),
        ("Form Detected", "Yes" if result.form_detected else "No"),
        ("Overall Confidence", f"{result.overall_confidence:.1%}"),
        ("Processed At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for i, (k, v) in enumerate(info_rows, start=2):
        ws1[f"A{i}"] = k
        ws1[f"B{i}"] = v
        ws1[f"A{i}"].border = thin
        ws1[f"B{i}"].border = thin

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 55

    # ------------------------------------------------------------------
    # Sheet 2 - Employee List
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Employee List", 1)
    headers = [
        "No", "Full Name", "Staff ID", "Department",
        "Attendance (hours)", "Exam Pass", "Exam Fail",
        "Discipline", "Course Result", "Certificate No", "Remark", "Confidence"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    for idx, emp in enumerate(result.employees, start=2):
        values = [
            emp.no,
            emp.full_name,
            emp.staff_id,
            emp.department,
            emp.attendance_hours,
            emp.exam_pass,
            emp.exam_fail,
            emp.discipline_status,
            emp.course_result,
            emp.certificate_no,
            "; ".join(emp.issues) if emp.issues else emp.remark,
            round(emp.confidence, 2),
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(idx, col, val)
            cell.border = thin
            if col == 9:  # Course Result
                if "Completed" in str(val):
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            if col == 12 and emp.confidence < 0.8:
                cell.fill = yellow_fill

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["K"].width = 40

    # ------------------------------------------------------------------
    # Sheet 3 - Error / Review
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Error Review", 2)
    ws3["A1"] = "TYPE"
    ws3["B1"] = "FIELD / STAFF"
    ws3["C1"] = "DETAIL"
    ws3["D1"] = "CONFIDENCE"
    for col in range(1, 5):
        cell = ws3.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

    row = 2
    # Global issues
    for iss in result.issues:
        ws3.cell(row, 1, "GLOBAL").border = thin
        ws3.cell(row, 2, "-").border = thin
        ws3.cell(row, 3, iss).border = thin
        ws3.cell(row, 4, result.overall_confidence).border = thin
        if "low" in iss.lower() or "missing" in iss.lower():
            ws3.cell(row, 3).fill = yellow_fill
        row += 1

    # Employee issues
    for emp in result.employees:
        for iss in emp.issues:
            ws3.cell(row, 1, "EMPLOYEE").border = thin
            ws3.cell(row, 2, f"{emp.staff_id} - {emp.full_name}").border = thin
            ws3.cell(row, 3, iss).border = thin
            ws3.cell(row, 4, emp.confidence).border = thin
            ws3.cell(row, 3).fill = yellow_fill
            row += 1

    if row == 2:
        ws3.cell(2, 1, "OK")
        ws3.cell(2, 3, "No issues found")

    for col, w in enumerate([14, 30, 70, 12], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # Save
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"export_{doc_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = export_dir / filename
    wb.save(filepath)
    return filepath
